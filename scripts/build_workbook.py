#!/usr/bin/env python3
"""Build Talero_NTConnect_Sheets_DB.xlsx from data/seed.json.

The workbook is the *datastore* for the mock when DATA_BACKEND=sheets: each NT Connect
resource group is one tab, row 1 = field headers (must match the JSON keys the server
reads/writes). It also carries a docs/README tab + the issued API keys so it reads as a
legit NT Connect API reference when imported to Google Sheets.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SEED = json.load(open(os.path.join(ROOT, "data", "seed.json")))
OUT = os.path.join(ROOT, "Talero_NTConnect_Sheets_DB.xlsx")

NAVY = "0B1F3A"
BLUE = "12407A"
LIGHT = "E8EEF6"
ACCENT = "1F7A4D"
GREY = "5A6472"

hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=BLUE)
title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=20)
title_fill = PatternFill("solid", fgColor=NAVY)
sub_font = Font(name="Calibri", color="FFFFFF", size=11)
key_font = Font(name="Consolas", color=NAVY, size=11, bold=True)
thin = Side(style="thin", color="C7D0DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Tabs that are real data tables (order matters for readability).
DATA_TABS = ["ApiKeys", "OrgUsers", "Accounts", "CashBalances", "Positions",
             "Orders", "Funds", "CustomerApplications", "RiskSettings",
             "Contracts", "Fees", "Alerts"]

ENDPOINTS = [
    ("Get Timestamp", "GET", "/v1/timestamp", "Server time (token-free)."),
    ("Authentication", "POST", "/v1/auth/accesstokenrequest", "Exchange cid + sec (API key) for a bearer access token."),
    ("Authentication", "POST", "/v1/auth/renewaccesstoken", "Refresh an unexpired access token."),
    ("OIDC User Info", "GET", "/v1/auth/me", "Profile of the authenticated org user."),
    ("Users", "GET", "/v1/user/list", "All org users (traders + admins)."),
    ("Users", "GET", "/v1/user/item?id=", "Single org user by id."),
    ("Users", "POST", "/v1/user/create", "Create one org user."),
    ("Users", "POST", "/v1/user/createbulk", "Bulk-create org users."),
    ("Accounting", "GET", "/v1/account/list", "All brokerage + sim accounts."),
    ("Accounting", "GET", "/v1/account/item?id=", "Single account by id."),
    ("Accounting", "GET", "/v1/account/deps?masterid=", "Accounts under a user/master."),
    ("Accounting", "POST", "/v1/account/create", "Create an account (org/AOP). Persists; retrievable immediately."),
    ("Accounting", "POST", "/v1/account/createbulk", "Bulk-create accounts."),
    ("Accounting", "GET", "/v1/cashBalance/list", "All cash-balance snapshots."),
    ("Accounting", "GET", "/v1/cashBalance/getcashbalancesnapshot?accountId=", "Latest snapshot: netLiq, openPnL, margins, availableForTrading."),
    ("Positions", "GET", "/v1/position/list?accountId=", "Open positions (optionally per account)."),
    ("Orders", "GET", "/v1/order/list?accountId=", "Orders (optionally per account)."),
    ("Funds", "GET", "/v1/funds/list?accountId=", "Deposits + withdrawals."),
    ("Funds", "POST", "/v1/funds/deposit", "Record a deposit (destination of record = NTC)."),
    ("Funds", "POST", "/v1/funds/withdraw", "Submit a withdrawal (return-to-originator, pending review)."),
    ("Risks", "GET", "/v1/risk/list?accountId=", "Pre/post-trade risk profiles."),
    ("Risks", "POST", "/v1/risk/apply", "Apply/update a risk profile to an account."),
    ("Risks", "POST", "/v1/risk/halt", "Halt/resume per-account or org-wide (kill switch)."),
    ("Customer Applications", "GET", "/v1/customerApplication/list", "AOP applications + statuses."),
    ("Customer Applications", "GET", "/v1/customerApplication/item?id=", "Single application."),
    ("Customer Applications", "POST", "/v1/customerApplication/create", "Submit a new AOP application."),
    ("Customer Applications", "POST", "/v1/customerApplication/approve", "Approve + issue an account."),
    ("Contract Library", "GET", "/v1/contract/list", "Tradable futures contracts."),
    ("Contract Library", "GET", "/v1/product/list", "Products behind the contracts."),
    ("Fees", "GET", "/v1/fee/list?accountId=", "Market-data + commission fees."),
    ("Alerts", "GET", "/v1/alert/list?accountId=", "Compliance / funding / onboarding alerts."),
    ("Talero App (not NT Connect)", "POST", "/app/register", "Customer self-registration for the Talero frontend."),
    ("Talero App (not NT Connect)", "POST", "/app/login", "Customer login; returns session + their accounts."),
    ("Talero App (not NT Connect)", "GET", "/app/me", "Authenticated customer profile + accounts."),
]

wb = Workbook()

# ---------------- READ ME ----------------
ws = wb.active
ws.title = "READ ME"
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1")
ws["A1"] = "NinjaTrader Connect — Partner API (Talero sandbox datastore)"
ws["A1"].font = title_font; ws["A1"].fill = title_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 38
meta = SEED.get("_meta", {})
lines = [
    ("Platform", "NT Connect is the Tradovate-based partner API. Bearer-token auth, relaxed REST (/v1/), JSON."),
    ("Environments", "live: live.tradovateapi.com  |  demo: demo.tradovateapi.com  |  market data: md.tradovateapi.com  |  staging: *.staging.ninjatrader.dev"),
    ("Access", "Org Admin credentials + API Key (sec) + CID. See the 'ApiKeys' tab."),
    ("Resource groups", "Get Timestamp - OIDC User Info - Users - Authentication - Accounting - Risks - Funds - Configuration - Alerts - Customer Applications - Orders - Contract Library - Personal Info - Fees - Positions"),
    ("How this workbook works", "Each tab below is one resource group / table. Row 1 = field names. The Talero mock server reads & writes these tabs live when DATA_BACKEND=sheets."),
    ("Endpoints", "See the 'Endpoints' tab for the full request surface this datastore serves."),
    ("SANDBOX NOTICE", meta.get("note", "Internal sandbox clone for development. Not affiliated with NinjaTrader; not a production brokerage backend.")),
]
r = 3
for k, v in lines:
    ws[f"A{r}"] = k; ws[f"A{r}"].font = Font(bold=True, color=BLUE)
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"] = v; ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30 if len(v) < 110 else 46
    r += 1
ws.column_dimensions["A"].width = 22
for c in "BCDEF": ws.column_dimensions[c].width = 22
ws.freeze_panes = "A3"

# ---------------- Endpoints ----------------
we = wb.create_sheet("Endpoints")
we.sheet_view.showGridLines = False
heads = ["Resource group", "Method", "Path", "Description"]
for j, h in enumerate(heads, 1):
    c = we.cell(1, j, h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
    c.alignment = Alignment(vertical="center")
for i, row in enumerate(ENDPOINTS, 2):
    for j, val in enumerate(row, 1):
        c = we.cell(i, j, val); c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=(j == 4))
        if j == 2:
            c.font = Font(bold=True, color=(ACCENT if val == "GET" else BLUE))
        if j == 3:
            c.font = Font(name="Consolas", color=NAVY)
we.column_dimensions["A"].width = 26
we.column_dimensions["B"].width = 9
we.column_dimensions["C"].width = 52
we.column_dimensions["D"].width = 64
we.freeze_panes = "A2"

def write_table(name, rows):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    if not rows:
        ws["A1"] = "(no rows)"; return
    # header = union of keys, first-row order preserved
    headers = list(rows[0].keys())
    for row in rows:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(vertical="center")
    for i, row in enumerate(rows, 2):
        for j, h in enumerate(headers, 1):
            v = row.get(h, "")
            if isinstance(v, (dict, list)):
                v = json.dumps(v, separators=(",", ":"))
            elif v is None:
                v = ""
            c = ws.cell(i, j, v); c.border = border
    # widths
    for j, h in enumerate(headers, 1):
        maxlen = max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows])
        ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 2, 10), 42)
    ws.freeze_panes = "A2"
    # special highlight for the API key secret column
    if name == "ApiKeys" and "apiKey" in headers:
        col = headers.index("apiKey") + 1
        for i in range(2, len(rows) + 2):
            ws.cell(i, col).font = key_font

for tab in DATA_TABS:
    write_table(tab, SEED.get(tab, []))

wb.save(OUT)
print("wrote", OUT)
print("tabs:", ", ".join([ws.title for ws in wb.worksheets]))
